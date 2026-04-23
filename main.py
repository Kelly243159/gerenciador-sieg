from fasthtml.common import *  # FastHTML + Starlette + HTMX helpers
import pandas as pd
from datetime import datetime
from unicodedata import normalize
import re, io, os, smtplib, hashlib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from starlette.responses import StreamingResponse, RedirectResponse
from starlette.middleware import Middleware
from starlette.middleware.sessions import SessionMiddleware

# -------------------------------------------------
# Configurações de E-mail (Modificável via variáveis de ambiente)
# -------------------------------------------------
SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
EMAIL_USER = os.getenv("EMAIL_USER", "")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD", "")
USE_TLS = os.getenv("USE_TLS", "True").lower() == "true"

# -------------------------------------------------
# Credenciais de Login (CONFIGURE VIA VARIÁVEIS DE AMBIENTE!)
# -------------------------------------------------
LOGIN_USER = os.getenv("LOGIN_USER", "mvtec2026")
LOGIN_PASSWORD = os.getenv("LOGIN_PASSWORD", "MV@@2026")
SESSION_SECRET = os.getenv("SESSION_SECRET", hashlib.sha256(b"mv-contabilidade-secret-2026").hexdigest())


# -------------------------------------------------
# Helpers
# -------------------------------------------------
def _only_digits(s: str) -> str:
    return re.sub(r"\D+", "", str(s) if s is not None else "")


def _norm(s: str) -> str:
    s = "" if s is None else str(s)
    s = normalize("NFKD", s).encode("ASCII", "ignore").decode("ASCII")
    return re.sub(r"[^a-z0-9 ]+", "", s.strip().lower())


def _pick_col(df, candidates):
    m = {_norm(c): c for c in df.columns}
    for c in candidates:
        k = _norm(c)
        if k in m:
            return m[k]
    return None


def _status(venc_dt, today=None):
    if today is None:
        today = datetime.today()
    if pd.isna(venc_dt):
        return "Sem data"
    d = (venc_dt - today).days
    if d < 0:
        return "Vencido"
    if d <= 30:
        return "A vencer"
    return "No prazo"


def gerar_relatorio(df_sieg, df_cert):
    cnpj_sieg = _pick_col(df_sieg, [
        "cpf_cnpj", "cpfcnpj", "cnpj", "cpf cnpj", "cpf/cnpj"
    ])
    cnpj_cert = _pick_col(df_cert, [
        "cnpj cpf", "cpf_cnpj", "cpfcnpj", "cnpj", "cpf cnpj", "cpf/cnpj"
    ])
    if not cnpj_sieg or not cnpj_cert:
        raise ValueError(
            f"Coluna CPF/CNPJ não encontrada.\n"
            f"  SIEG cols: {list(df_sieg.columns)}\n"
            f"  Cert cols: {list(df_cert.columns)}"
        )

    col_resp  = _pick_col(df_sieg, ["responsavel", "responsável", "nome do responsavel"])
    col_emp   = _pick_col(df_sieg, ["empresa", "razao social", "razão social",
                                     "cliente", "nome do cliente", "nome fantasia"])
    col_email = _pick_col(df_sieg, ["email", "e-mail", "email1", "e-mail1"])

    col_venc = _pick_col(df_cert, [
        "vencimento cert titular",
        "vencimento cert procurador",
        "vencimento", "validade",
        "data de vencimento", "data vencimento",
    ])

    df_sieg["_CPF_CNPJ_"] = df_sieg[cnpj_sieg].map(_only_digits)
    df_cert["_CPF_CNPJ_"] = df_cert[cnpj_cert].map(_only_digits)

    keep = ["_CPF_CNPJ_"] + ([col_venc] if col_venc else [])
    df_cert_small = df_cert[keep].drop_duplicates(subset=["_CPF_CNPJ_"]).copy()

    merged = pd.merge(df_sieg, df_cert_small, on="_CPF_CNPJ_", how="left")

    out = pd.DataFrame()
    out["Responsavel"] = merged[col_resp].fillna("") if col_resp else ""
    out["Empresa"]     = merged[col_emp].fillna("")  if col_emp  else ""
    out["Email"]       = merged[col_email].fillna("") if col_email else ""
    out["CPF_CNPJ"]    = merged["_CPF_CNPJ_"]

    if col_venc:
        venc = pd.to_datetime(merged[col_venc], errors="coerce", dayfirst=True)
        out["Vencimento"] = venc.dt.strftime("%d/%m/%Y").fillna("")
        out["Status"]     = [_status(d) for d in venc]
    else:
        out["Vencimento"] = ""
        out["Status"]     = "Sem data"

    return out


# -------------------------------------------------
# Excel estilizado
# -------------------------------------------------
def make_excel_bytes(df: pd.DataFrame, sheet_name="Relatorio") -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        header_fill  = PatternFill(start_color="1a1f36", end_color="1a1f36", fill_type="solid")
        header_font  = Font(color="FFFFFF", bold=True)
        red_fill     = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        yellow_fill  = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        blue_fill    = PatternFill(start_color="CCE0FF", end_color="CCE0FF", fill_type="solid")
        gray_fill    = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = max(14, len(str(df.columns[col - 1])) + 4)

        status_col_idx = df.columns.get_loc("Status") + 1
        for row in range(2, len(df) + 2):
            sv = ws.cell(row=row, column=status_col_idx).value
            fill = (red_fill    if sv == "Vencido"  else
                    yellow_fill if sv == "A vencer" else
                    blue_fill   if sv == "No prazo" else gray_fill)
            for col in range(1, len(df.columns) + 1):
                ws.cell(row=row, column=col).fill      = fill
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="left", vertical="center")

    buf.seek(0)
    return buf.getvalue()


# -------------------------------------------------
# E-mail via SMTP
# -------------------------------------------------
def enviar_email_smtp(destinatario: str, assunto: str, body_html: str):
    try:
        if not all([SMTP_SERVER, EMAIL_USER, EMAIL_PASSWORD]):
            return False, "Configuração de e-mail incompleta. Configure SMTP_SERVER, EMAIL_USER e EMAIL_PASSWORD."

        msg = MIMEMultipart()
        msg['From'] = EMAIL_USER
        msg['To'] = destinatario
        msg['Subject'] = assunto
        msg.attach(MIMEText(body_html, 'html'))

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            if USE_TLS:
                server.starttls()
            server.login(EMAIL_USER, EMAIL_PASSWORD)
            server.send_message(msg)

        return True, "E-mail enviado com sucesso"
    except Exception as e:
        return False, f"Falha ao enviar e-mail: {str(e)}"


def corpo_email_vencido(empresa: str, vencimento: str) -> str:
    return f"""
    <div style="font-family:Segoe UI,Arial,sans-serif;font-size:14px;color:#111">
      <p>Prezado(a) Cliente,</p>
      <p>Identificamos que o certificado digital da empresa <strong>{empresa}</strong>
      encontra-se <strong style='color:#b91c1c'>VENCIDO</strong> desde <strong>{vencimento}</strong>.</p>
      <p style="margin-top:18px"><strong>⚠️ É urgente regularizar</strong> para evitar:
      <ul>
        <li>Bloqueio de acesso a sistemas governamentais (Receita, SEFAZ, Prefeituras)</li>
        <li>Impossibilidade de emissão de NF-e/NFS-e</li>
        <li>Interrupção de procurações e transmissões</li>
        <li>Multas e outras penalidades</li>
      </ul>
      </p>
      <div style="margin:18px 0;padding:14px 16px;border:1px solid #e5e7eb;border-radius:10px;background:#f8fafc">
        <p style="margin:0 0 6px 0"><strong>💠 MV CONTABILIDADE | CERTIFICADORA DIGITAL</strong></p>
        <ul style="margin:8px 0 0 18px">
          <li>Emissão/renovação de certificados <strong>A1</strong> e <strong>A3</strong></li>
          <li>Atendimento rápido e acompanhamento completo</li>
          <li>Suporte técnico para instalação e uso</li>
          <li>Condições especiais para clientes MV</li>
        </ul>
      </div>
      <p>Ficamos à disposição para proceder com a renovação imediatamente.</p>
      <p>
        📞 <strong>Contato MV:</strong> (41) 99673-1918<br>
        ✉️ <strong>Entre em contato com o seu contador responsável</strong>
      </p>
      <p>Atenciosamente,<br>
      <strong>Equipe MV Contabilidade</strong></p>
    </div>
    """


def corpo_email_a_vencer(empresa: str, vencimento: str) -> str:
    return f"""
    <div style="font-family:Segoe UI,Arial,sans-serif;font-size:14px;color:#111">
      <p>Prezado(a) Cliente,</p>
      <p>Identificamos que o certificado digital da empresa <strong>{empresa}</strong>
      <strong style='color:#f59e0b'>VENCE EM BREVE</strong> no dia <strong>{vencimento}</strong>.</p>
      <p style="margin-top:18px"><strong>📅 É importante renovar antecipadamente</strong> para evitar:
      <ul>
        <li>Bloqueio de acesso a sistemas governamentais</li>
        <li>Impossibilidade de emissão de documentos fiscais</li>
        <li>Interrupção das atividades da empresa</li>
      </ul>
      </p>
      <div style="margin:18px 0;padding:14px 16px;border:1px solid #e5e7eb;border-radius:10px;background:#f8fafc">
        <p style="margin:0 0 6px 0"><strong>💠 MV CONTABILIDADE | CERTIFICADORA DIGITAL</strong></p>
        <ul style="margin:8px 0 0 18px">
          <li>Emissão/renovação de certificados <strong>A1</strong> e <strong>A3</strong></li>
          <li>Atendimento rápido e acompanhamento completo</li>
          <li>Suporte técnico para instalação e uso</li>
          <li>Condições especiais para clientes MV</li>
        </ul>
      </div>
      <p>Entre em contato conosco para renovar seu certificado.</p>
      <p>
        📞 <strong>Contato MV:</strong> (41) 99673-1918<br>
        ✉️ <strong>Entre em contato com o seu contador responsável</strong>
      </p>
      <p>Atenciosamente,<br>
      <strong>Equipe MV Contabilidade</strong></p>
    </div>
    """


# -------------------------------------------------
# FastHTML app (com SessionMiddleware)
# -------------------------------------------------
app, rt = fast_app(
    middleware=[Middleware(SessionMiddleware, secret_key=SESSION_SECRET)]
)
current_data = {}


# -------------------------------------------------
# Helper: verificar autenticação
# -------------------------------------------------
def is_authenticated(request):
    return request.session.get("authenticated") is True


# =====================================================================
# CSS — LOGIN
# =====================================================================
def login_css():
    return Style("""
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:ital,opsz,wght@0,9..40,300;0,9..40,400;0,9..40,500;0,9..40,600;0,9..40,700;0,9..40,800;1,9..40,400&family=Space+Mono:wght@400;700&display=swap');

*{box-sizing:border-box;margin:0;padding:0}

body{
    min-height:100vh;
    background:#08090e;
    font-family:'DM Sans',system-ui,sans-serif;
    overflow-x:hidden;
}

.login-canvas{
    min-height:100vh;
    display:grid;
    grid-template-columns:1fr 1fr;
    position:relative;
}
@media(max-width:860px){
    .login-canvas{grid-template-columns:1fr}
    .login-brand-side{display:none}
}

.login-brand-side{
    background:linear-gradient(160deg,#0d1117 0%,#0c1528 40%,#0a1a3d 100%);
    display:flex;flex-direction:column;justify-content:center;align-items:center;
    padding:60px 48px;position:relative;overflow:hidden;
}
.login-brand-side::before{
    content:"";position:absolute;width:500px;height:500px;border-radius:50%;
    background:radial-gradient(circle,rgba(59,130,246,.1) 0%,transparent 70%);
    top:-100px;right:-100px;pointer-events:none;
}
.login-brand-side::after{
    content:"";position:absolute;width:400px;height:400px;border-radius:50%;
    background:radial-gradient(circle,rgba(56,189,248,.06) 0%,transparent 70%);
    bottom:-80px;left:-80px;pointer-events:none;
}

.brand-logo-box{
    width:96px;height:96px;border-radius:24px;
    background:linear-gradient(145deg,#3b82f6 0%,#2563eb 50%,#1d4ed8 100%);
    display:flex;align-items:center;justify-content:center;
    font-family:'Space Mono',monospace;font-weight:700;font-size:32px;color:#fff;
    letter-spacing:-.04em;
    box-shadow:0 24px 64px rgba(59,130,246,.25),inset 0 2px 0 rgba(255,255,255,.2);
    margin-bottom:32px;position:relative;
}
.brand-logo-box::after{
    content:"";position:absolute;inset:0;border-radius:24px;
    background:linear-gradient(135deg,rgba(255,255,255,.15) 0%,transparent 50%);pointer-events:none;
}

.brand-name{font-family:'Space Mono',monospace;font-size:1.1rem;font-weight:700;color:#eff6ff;
    letter-spacing:.15em;text-transform:uppercase;margin-bottom:12px;}
.brand-tagline{font-size:.95rem;color:#94a3b8;text-align:center;line-height:1.6;max-width:320px;}
.brand-decorline{width:48px;height:3px;border-radius:3px;
    background:linear-gradient(90deg,#3b82f6,#38bdf8);margin:28px 0 24px;}
.brand-features{list-style:none;padding:0;margin:0;}
.brand-features li{color:#64748b;font-size:.82rem;padding:6px 0;display:flex;align-items:center;gap:10px;}
.brand-features li::before{content:"";width:6px;height:6px;border-radius:50%;background:#3b82f6;flex-shrink:0;}

.login-form-side{display:flex;align-items:center;justify-content:center;padding:48px 36px;background:#08090e;position:relative;}
.login-box{width:100%;max-width:380px;}
.login-box-header{margin-bottom:36px;}
.login-box-title{font-size:1.6rem;font-weight:800;color:#f1f5f9;letter-spacing:-.03em;line-height:1.2;}
.login-box-sub{font-size:.85rem;color:#64748b;margin-top:6px;line-height:1.5;}

.field{margin-bottom:22px}
.field-label{display:block;font-size:.7rem;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.1em;margin-bottom:8px;}
.field-input{
    width:100%;padding:14px 16px;background:rgba(255,255,255,.04);
    border:1px solid rgba(255,255,255,.08);border-radius:10px;color:#e2e8f0;
    font-family:'DM Sans',sans-serif;font-size:.92rem;font-weight:500;outline:none;transition:all .2s;
}
.field-input::placeholder{color:#334155;font-weight:400}
.field-input:focus{border-color:#3b82f6;box-shadow:0 0 0 3px rgba(59,130,246,.15);background:rgba(255,255,255,.06);}

.login-btn{
    width:100%;padding:15px;margin-top:8px;
    background:linear-gradient(135deg,#3b82f6,#2563eb);color:#fff;border:none;border-radius:10px;
    font-family:'DM Sans',sans-serif;font-size:.92rem;font-weight:700;cursor:pointer;transition:all .2s;
    box-shadow:0 8px 32px rgba(59,130,246,.25);letter-spacing:.01em;
}
.login-btn:hover{transform:translateY(-2px);box-shadow:0 12px 40px rgba(59,130,246,.35);}
.login-btn:active{transform:translateY(0)}

.login-error{
    background:rgba(239,68,68,.06);border:1px solid rgba(239,68,68,.15);
    border-radius:10px;padding:12px 16px;margin-bottom:22px;
    display:flex;align-items:center;gap:10px;animation:shake .35s ease-in-out;
}
.login-error-icon{font-size:1rem;flex-shrink:0}
.login-error-text{color:#fca5a5;font-size:.82rem;font-weight:500;line-height:1.4}
@keyframes shake{0%,100%{transform:translateX(0)}25%{transform:translateX(-5px)}75%{transform:translateX(5px)}}

.login-footer{text-align:center;margin-top:28px;font-size:.72rem;color:#1e293b;}
""")


def login_page(error=None):
    return (
        login_css(),
        Main(
            Div(
                Div(
                    Div("MV", cls="brand-logo-box"),
                    Div("CERTIFICADORA MV", cls="brand-name"),
                    Div("Sistema profissional de gerenciamento e notificação de certificados digitais", cls="brand-tagline"),
                    Div(cls="brand-decorline"),
                    Ul(
                        Li("Controle de vencimentos em tempo real"),
                        Li("Envio automático de notificações"),
                        Li("Relatórios exportáveis em Excel"),
                        Li("Dashboard completo por status"),
                        cls="brand-features"
                    ),
                    cls="login-brand-side"
                ),
                Div(
                    Div(
                        Div(
                            Div("Acesse sua conta", cls="login-box-title"),
                            Div("Entre com suas credenciais para acessar o painel", cls="login-box-sub"),
                            cls="login-box-header"
                        ),
                        Div(
                            Span("⚠", cls="login-error-icon"),
                            Span("Credenciais incorretas. Verifique e tente novamente.", cls="login-error-text"),
                            cls="login-error"
                        ) if error else "",
                        Form(
                            Div(
                                Label("Usuário", cls="field-label"),
                                Input(type="text", name="username", placeholder="seu.usuario",
                                      required=True, autocomplete="username", cls="field-input"),
                                cls="field"
                            ),
                            Div(
                                Label("Senha", cls="field-label"),
                                Input(type="password", name="password", placeholder="••••••••",
                                      required=True, autocomplete="current-password", cls="field-input"),
                                cls="field"
                            ),
                            Button("Entrar", type="submit", cls="login-btn"),
                            method="post", action="/login"
                        ),
                        Div("MV Contabilidade © 2026 — Acesso restrito", cls="login-footer"),
                        cls="login-box"
                    ),
                    cls="login-form-side"
                ),
                cls="login-canvas"
            )
        )
    )


# =====================================================================
# CSS — APP GLOBAL (Upload + Dashboard)
# =====================================================================
def global_css():
    return Style("""
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:ital,opsz,wght@0,9..40,300;0,9..40,400;0,9..40,500;0,9..40,600;0,9..40,700;0,9..40,800;1,9..40,400&family=Space+Mono:wght@400;700&display=swap');

:root{
    --bg:#08090e;--surface:#0f1117;--surface-2:#161922;
    --border:rgba(255,255,255,.06);--border-hover:rgba(255,255,255,.12);
    --text:#f1f5f9;--text-2:#94a3b8;--text-3:#475569;
    --blue:#3b82f6;--blue-dark:#2563eb;--blue-deeper:#1d4ed8;
    --red:#ef4444;--red-dark:#dc2626;--amber:#f59e0b;--amber-dark:#d97706;--sky:#38bdf8;
    --radius:14px;--radius-lg:20px;
    --dash-bg:#f4f7fb;--card-bg:#ffffff;--card-border:#e5e9f0;
    --card-shadow:0 4px 24px rgba(15,23,42,.06);
    --dark-text:#0f172a;--dark-text-2:#475569;
}

*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--text);font-family:'DM Sans',system-ui,sans-serif;min-height:100vh;-webkit-font-smoothing:antialiased;}
.container{max-width:1280px;margin:0 auto;padding:32px 24px}

.topbar{display:flex;align-items:center;justify-content:space-between;padding:16px 24px;background:var(--surface);border-bottom:1px solid var(--border);position:sticky;top:0;z-index:100;backdrop-filter:blur(12px);}
.topbar-left{display:flex;align-items:center;gap:12px}
.topbar-logo{width:36px;height:36px;border-radius:10px;background:linear-gradient(145deg,#3b82f6,#2563eb);display:flex;align-items:center;justify-content:center;font-family:'Space Mono',monospace;font-weight:700;font-size:13px;color:#fff;box-shadow:0 4px 16px rgba(59,130,246,.2);}
.topbar-title{font-family:'Space Mono',monospace;font-size:.78rem;font-weight:700;color:#f1f5f9;letter-spacing:.1em;text-transform:uppercase;}
.btn-logout{background:transparent;color:#64748b;border:1px solid var(--border);border-radius:8px;padding:8px 16px;font-size:.78rem;font-weight:600;cursor:pointer;transition:all .2s;text-decoration:none;font-family:'DM Sans',sans-serif;display:inline-flex;align-items:center;gap:6px;}
.btn-logout:hover{background:rgba(239,68,68,.08);border-color:rgba(239,68,68,.2);color:#fca5a5;}

.upload-hero{text-align:center;padding:48px 0 16px;}
.upload-logo{width:72px;height:72px;border-radius:20px;background:linear-gradient(145deg,#3b82f6,#2563eb,#1d4ed8);display:inline-flex;align-items:center;justify-content:center;font-family:'Space Mono',monospace;font-weight:700;font-size:26px;color:#fff;box-shadow:0 20px 48px rgba(59,130,246,.2);margin-bottom:20px;position:relative;animation:float-subtle 4s ease-in-out infinite;}
@keyframes float-subtle{0%,100%{transform:translateY(0)}50%{transform:translateY(-6px)}}
.upload-hero h1{font-size:1.8rem;font-weight:800;color:#f1f5f9;letter-spacing:-.03em;line-height:1.15;margin-bottom:6px;}
.upload-hero p{color:#64748b;font-size:.9rem;line-height:1.5;}

.upload-card{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-lg);padding:32px;max-width:640px;margin:24px auto 0;box-shadow:0 24px 64px rgba(0,0,0,.3);}
.upload-field{margin-bottom:20px}
.upload-label{display:flex;align-items:center;gap:8px;font-size:.8rem;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:.06em;margin-bottom:10px;}
.upload-label .dot{width:7px;height:7px;border-radius:50%;background:var(--blue);flex-shrink:0;}

.filebox{width:100%;padding:18px 20px;border:1.5px dashed rgba(59,130,246,.3);border-radius:12px;background:rgba(59,130,246,.03);color:var(--text);cursor:pointer;transition:all .2s;font-family:'DM Sans',sans-serif;font-size:.88rem;}
.filebox:hover{border-color:rgba(59,130,246,.5);background:rgba(59,130,246,.06);}
.filebox::file-selector-button{background:var(--blue);color:#fff;border:none;border-radius:8px;padding:8px 16px;font-weight:600;cursor:pointer;margin-right:14px;font-family:'DM Sans',sans-serif;font-size:.82rem;transition:background .2s;}
.filebox::file-selector-button:hover{background:var(--blue-dark)}

.btn-process{width:100%;padding:16px;margin-top:8px;background:linear-gradient(135deg,#3b82f6,#2563eb);color:#fff;border:none;border-radius:12px;font-family:'DM Sans',sans-serif;font-size:.95rem;font-weight:700;cursor:pointer;transition:all .2s;box-shadow:0 8px 32px rgba(59,130,246,.2);letter-spacing:.01em;}
.btn-process:hover{transform:translateY(-2px);box-shadow:0 12px 40px rgba(59,130,246,.3);}

.dash-shell{background:var(--dash-bg);border-radius:var(--radius-lg);padding:28px;min-height:80vh;}
.dash-header{background:var(--card-bg);border:1px solid var(--card-border);border-radius:16px;box-shadow:var(--card-shadow);padding:20px 24px;margin-bottom:24px;display:flex;flex-wrap:wrap;align-items:center;gap:16px;}
.dash-header-icon{width:44px;height:44px;border-radius:12px;background:linear-gradient(145deg,#3b82f6,#2563eb);display:flex;align-items:center;justify-content:center;font-family:'Space Mono',monospace;font-weight:700;font-size:15px;color:#fff;box-shadow:0 8px 24px rgba(59,130,246,.2);flex-shrink:0;}
.dash-header-text{flex:1;min-width:200px}
.dash-header-title{font-size:1rem;font-weight:700;color:var(--dark-text);display:flex;flex-wrap:wrap;align-items:center;gap:10px;}
.dash-badge{font-size:.65rem;font-weight:700;padding:4px 10px;border-radius:999px;background:linear-gradient(135deg,#3b82f6,#2563eb);color:#fff;text-transform:uppercase;letter-spacing:.05em;}
.dash-header-sub{font-size:.82rem;color:var(--dark-text-2);margin-top:2px;}

.stats-row{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:14px;margin-bottom:24px;}
.stat-card{background:var(--card-bg);border:1px solid var(--card-border);border-radius:14px;padding:18px 20px;box-shadow:var(--card-shadow);position:relative;overflow:hidden;transition:transform .15s,box-shadow .15s;}
.stat-card:hover{transform:translateY(-2px);box-shadow:0 8px 32px rgba(15,23,42,.1);}
.stat-card::before{content:"";position:absolute;top:0;left:0;right:0;height:3px;border-radius:3px 3px 0 0;}
.stat-card[data-accent="total"]::before{background:linear-gradient(90deg,#6366f1,#818cf8)}
.stat-card[data-accent="red"]::before{background:linear-gradient(90deg,#ef4444,#f87171)}
.stat-card[data-accent="amber"]::before{background:linear-gradient(90deg,#f59e0b,#fbbf24)}
.stat-card[data-accent="blue"]::before{background:linear-gradient(90deg,#3b82f6,#60a5fa)}
.stat-card[data-accent="gray"]::before{background:linear-gradient(90deg,#64748b,#94a3b8)}
.stat-label{font-size:.72rem;font-weight:600;color:var(--dark-text-2);text-transform:uppercase;letter-spacing:.06em}
.stat-num{font-size:1.8rem;font-weight:800;color:var(--dark-text);line-height:1.2;margin-top:4px;font-family:'Space Mono',monospace;}

.sec{background:var(--card-bg);border:1px solid var(--card-border);border-radius:16px;box-shadow:var(--card-shadow);margin-bottom:24px;overflow:hidden;}
.sec-header{padding:18px 24px;display:flex;flex-wrap:wrap;align-items:center;justify-content:space-between;gap:12px;border-bottom:1px solid var(--card-border);}
.sec-title{font-size:.95rem;font-weight:700;color:var(--dark-text)}
.sec-desc{font-size:.8rem;color:var(--dark-text-2);margin-top:2px}
.sec-body{padding:0}
.sec-body-padded{padding:20px 24px}

.btn{display:inline-flex;align-items:center;gap:6px;padding:10px 16px;border:none;border-radius:10px;font-family:'DM Sans',sans-serif;font-size:.8rem;font-weight:700;color:#fff;cursor:pointer;transition:all .2s;text-decoration:none;line-height:1.2;}
.btn:hover{transform:translateY(-1px)}
.btn-blue{background:linear-gradient(135deg,#3b82f6,#2563eb);box-shadow:0 4px 16px rgba(59,130,246,.2)}
.btn-red{background:linear-gradient(135deg,#ef4444,#dc2626);box-shadow:0 4px 16px rgba(239,68,68,.2)}
.btn-amber{background:linear-gradient(135deg,#f59e0b,#d97706);box-shadow:0 4px 16px rgba(245,158,11,.2)}

.tbl{width:100%;border-collapse:collapse;font-size:.85rem}
.tbl thead th{background:#f8fafc;color:var(--dark-text);font-weight:600;text-align:left;padding:12px 16px;border-bottom:1px solid #e5e9f0;font-size:.68rem;text-transform:uppercase;letter-spacing:.06em;position:sticky;top:0;z-index:2;}
.tbl tbody tr{border-bottom:1px solid #f1f5f9;transition:background .1s;}
.tbl tbody td{padding:12px 16px;color:var(--dark-text);vertical-align:middle;}
.tbl tbody tr.r-ok{background:#eff6ff}.tbl tbody tr.r-ok:hover{background:#dbeafe}
.tbl tbody tr.r-vencido{background:#fef2f2}.tbl tbody tr.r-vencido:hover{background:#fee2e2}
.tbl tbody tr.r-avencer{background:#fffbeb}.tbl tbody tr.r-avencer:hover{background:#fef3c7}
.tbl tbody tr.r-semdata{background:#f8fafc}.tbl tbody tr.r-semdata:hover{background:#f1f5f9}

.pill{display:inline-block;font-size:.66rem;font-weight:700;padding:4px 10px;border-radius:999px;color:#fff;text-transform:uppercase;letter-spacing:.04em;}
.pill-red{background:linear-gradient(135deg,#ef4444,#dc2626)}
.pill-amber{background:linear-gradient(135deg,#f59e0b,#d97706)}
.pill-blue{background:linear-gradient(135deg,#3b82f6,#2563eb)}
.pill-gray{background:linear-gradient(135deg,#64748b,#475569)}

.co-grid{display:grid;gap:14px;grid-template-columns:repeat(auto-fill,minmax(300px,1fr));padding:20px 24px;}
.co-card{background:var(--surface);color:#fff;border:1px solid var(--border);border-radius:14px;padding:18px 20px;box-shadow:0 8px 32px rgba(0,0,0,.3);transition:transform .15s,box-shadow .15s;}
.co-card:hover{transform:translateY(-2px);box-shadow:0 12px 40px rgba(0,0,0,.4);}
.co-top{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:10px}
.co-badge{padding:5px 10px;border-radius:8px;font-size:.68rem;font-weight:700;color:#fff;text-transform:uppercase;letter-spacing:.03em;}
.co-badge-red{background:linear-gradient(135deg,#ef4444,#b91c1c)}
.co-badge-amber{background:linear-gradient(135deg,#f59e0b,#b45309)}
.co-name{font-size:.95rem;font-weight:700;color:#f1f5f9;line-height:1.3}
.co-info{font-size:.82rem;color:#94a3b8;margin-top:4px;line-height:1.4}
.co-email{font-size:.78rem;color:#64748b;margin-top:8px}
.co-btn{padding:7px 12px;border:none;border-radius:8px;font-family:'DM Sans',sans-serif;font-size:.75rem;font-weight:700;color:#fff;cursor:pointer;transition:all .2s;}
.co-btn-red{background:linear-gradient(135deg,#ef4444,#dc2626)}
.co-btn-amber{background:linear-gradient(135deg,#f59e0b,#d97706)}
.co-btn:hover{transform:translateY(-1px)}
.inline-feedback{margin-top:10px;font-size:.78rem;font-weight:600}
.inline-feedback.ok{color:#3b82f6}
.inline-feedback.err{color:#ef4444}

.empty-state{padding:28px;text-align:center;font-size:.88rem;font-weight:600;border-radius:12px;}
.empty-ok{background:#eff6ff;color:#2563eb;border:1px solid #dbeafe}
.empty-neutral{background:#f8fafc;color:#64748b;border:1px solid #e5e9f0}

.mt-2{margin-top:14px}.mt-3{margin-top:20px}.text-center{text-align:center}
.sec table{width:100%;border-collapse:collapse;font-size:.84rem}
.sec th,.sec td{padding:12px 16px;border-bottom:1px solid #f1f5f9;color:var(--dark-text);text-align:left}
.sec th{background:#f8fafc;font-weight:600;font-size:.7rem;text-transform:uppercase;letter-spacing:.05em}
.tbl-scroll{max-height:560px;overflow:auto}
""")


def topbar(show_logout=True):
    return Div(
        Div(
            Div("MV", cls="topbar-logo"),
            Span("CERTIFICADORA MV", cls="topbar-title"),
            cls="topbar-left"
        ),
        A("Sair →", href="/logout", cls="btn-logout") if show_logout else "",
        cls="topbar"
    )


# =====================================================================
# TELA 1 — UPLOAD
# =====================================================================
def page():
    return (
        global_css(),
        Main(
            topbar(),
            Section(
                Div(
                    Div("MV", cls="upload-logo"),
                    H1("Gerenciamento de Certificados"),
                    P("Importe as planilhas SIEG e de Certificados para gerar o relatório completo"),
                    cls="upload-hero"
                ),
                Div(
                    Form(
                        Div(
                            Div(
                                Span(cls="dot"),
                                Span("Planilha SIEG — Clientes e Responsáveis"),
                                cls="upload-label"
                            ),
                            Input(type="file", name="file_sieg", accept=".xlsx,.xls",
                                  required=True, cls="filebox"),
                            cls="upload-field"
                        ),
                        Div(
                            Div(
                                Span(cls="dot"),
                                Span("Planilha de Certificados — Datas de Vencimento"),
                                cls="upload-label"
                            ),
                            Input(type="file", name="file_cert", accept=".xlsx,.xls",
                                  required=True, cls="filebox"),
                            cls="upload-field"
                        ),
                        Button("Processar Planilhas →", type="submit", cls="btn-process"),
                        method="post", action="/processar-upload", enctype="multipart/form-data"
                    ),
                    cls="upload-card"
                ),
                cls="container"
            ),
        )
    )


# -------------------------------------------------
# Rotas de Autenticação
# -------------------------------------------------
@app.get("/login")
def login_get(request):
    if is_authenticated(request):
        return RedirectResponse("/", status_code=303)
    return Titled("Login — CERTIFICADORA MV", login_page())


@app.post("/login")
async def login_post(request):
    form = await request.form()
    username = form.get("username", "").strip()
    password = form.get("password", "")

    if username == LOGIN_USER and password == LOGIN_PASSWORD:
        request.session["authenticated"] = True
        return RedirectResponse("/", status_code=303)
    else:
        return Titled("Login — CERTIFICADORA MV", login_page(error=True))


@app.get("/logout")
def logout(request):
    request.session.clear()
    return RedirectResponse("/login", status_code=303)


# -------------------------------------------------
# Rotas principais
# -------------------------------------------------
@app.get("/")
def index(request):
    if not is_authenticated(request):
        return RedirectResponse("/login", status_code=303)
    return Titled("CERTIFICADORA MV — Dashboard", page())


@app.post("/processar-upload")
async def processar_upload(request):
    if not is_authenticated(request):
        return RedirectResponse("/login", status_code=303)

    try:
        form = await request.form()
        file_sieg = form.get("file_sieg")
        file_cert = form.get("file_cert")
        if not file_sieg or not file_cert:
            return Div(Div(P("Selecione ambas as planilhas."), cls="upload-card"), cls="container")

        df_sieg = pd.read_excel(io.BytesIO(await file_sieg.read()), dtype=str)
        df_cert = pd.read_excel(io.BytesIO(await file_cert.read()), dtype=str)

        df = gerar_relatorio(df_sieg, df_cert)

        data_id = hashlib.md5(str(datetime.now()).encode()).hexdigest()[:8]
        current_data[data_id] = df
        current_data[f"{data_id}_excel"] = make_excel_bytes(df)

        total   = len(df)
        vencidos = (df['Status'] == 'Vencido').sum()
        avencer  = (df['Status'] == 'A vencer').sum()
        noprazo  = (df['Status'] == 'No prazo').sum()
        semdata  = (df['Status'] == 'Sem data').sum()

        linhas = []
        for idx, row in df.iterrows():
            st = row['Status']
            rcls = ("r-vencido" if st == "Vencido" else
                    "r-avencer" if st == "A vencer" else
                    "r-ok"      if st == "No prazo" else "r-semdata")
            pcls = ("pill-red"   if st == "Vencido" else
                    "pill-amber" if st == "A vencer" else
                    "pill-blue"  if st == "No prazo" else "pill-gray")
            linhas.append(
                Tr(
                    Td(row['Responsavel']),
                    Td(row['Empresa']),
                    Td(row['Email']),
                    Td(row['CPF_CNPJ']),
                    Td(row['Vencimento']),
                    Td(Span(st, cls=f"pill {pcls}")),
                    cls=rcls
                )
            )

        cards_vencidos = []
        for idx, row in df[df['Status'] == 'Vencido'].iterrows():
            email = (row.get('Email') or '').strip()
            empresa = row.get('Empresa', '') or '—'
            vencimento = row.get('Vencimento', '') or '—'
            cnpj_txt = row.get('CPF_CNPJ', '')
            cid = f"vencido_{data_id}_{idx}"

            cards_vencidos.append(
                Div(
                    Div(
                        Span("VENCIDO", cls="co-badge co-badge-red"),
                        Button(
                            "Enviar E-mail",
                            cls="co-btn co-btn-red",
                            hx_post=f"/envia-email-individual/{data_id}/vencido/{idx}",
                            hx_target=f"#res_{cid}",
                            hx_swap="innerHTML",
                        ) if email and '@' in email else
                        Span("Sem e-mail", style="color:#64748b;font-size:.75rem;font-weight:600"),
                        cls="co-top"
                    ),
                    Div(empresa, cls="co-name"),
                    Div(f"Vencimento: {vencimento}  ·  CNPJ: {cnpj_txt}", cls="co-info"),
                    Div(f"✉ {email}", cls="co-email") if email and '@' in email else "",
                    Div(id=f"res_{cid}", cls="inline-feedback"),
                    cls="co-card"
                )
            )

        cards_avencer = []
        for idx, row in df[df['Status'] == 'A vencer'].iterrows():
            email = (row.get('Email') or '').strip()
            empresa = row.get('Empresa', '') or '—'
            vencimento = row.get('Vencimento', '') or '—'
            cnpj_txt = row.get('CPF_CNPJ', '')
            cid = f"avencer_{data_id}_{idx}"

            cards_avencer.append(
                Div(
                    Div(
                        Span("A VENCER", cls="co-badge co-badge-amber"),
                        Button(
                            "Enviar E-mail",
                            cls="co-btn co-btn-amber",
                            hx_post=f"/envia-email-individual/{data_id}/avencer/{idx}",
                            hx_target=f"#res_{cid}",
                            hx_swap="innerHTML",
                        ) if email and '@' in email else
                        Span("Sem e-mail", style="color:#64748b;font-size:.75rem;font-weight:600"),
                        cls="co-top"
                    ),
                    Div(empresa, cls="co-name"),
                    Div(f"Vencimento: {vencimento}  ·  CNPJ: {cnpj_txt}", cls="co-info"),
                    Div(f"✉ {email}", cls="co-email") if email and '@' in email else "",
                    Div(id=f"res_{cid}", cls="inline-feedback"),
                    cls="co-card"
                )
            )

        return Div(
            global_css(),
            topbar(),
            Div(
                Div(
                    Div(
                        Div("MV", cls="dash-header-icon"),
                        Div(
                            Div(
                                "Painel de Certificados",
                                Span("MV CONTABILIDADE", cls="dash-badge"),
                                cls="dash-header-title"
                            ),
                            Div("Visão geral dos certificados digitais processados", cls="dash-header-sub"),
                            cls="dash-header-text"
                        ),
                        cls="dash-header"
                    ),
                    Div(
                        Div(Div("Total", cls="stat-label"), Div(str(total), cls="stat-num"), cls="stat-card", **{"data-accent": "total"}),
                        Div(Div("Vencidos", cls="stat-label"), Div(str(vencidos), cls="stat-num"), cls="stat-card", **{"data-accent": "red"}),
                        Div(Div("A vencer", cls="stat-label"), Div(str(avencer), cls="stat-num"), cls="stat-card", **{"data-accent": "amber"}),
                        Div(Div("No prazo", cls="stat-label"), Div(str(noprazo), cls="stat-num"), cls="stat-card", **{"data-accent": "blue"}),
                        Div(Div("Sem data", cls="stat-label"), Div(str(semdata), cls="stat-num"), cls="stat-card", **{"data-accent": "gray"}),
                        cls="stats-row"
                    ),
                    Div(
                        Div(
                            Div(Div("Relatório Completo", cls="sec-title"), Div("Todos os certificados com indicador de status", cls="sec-desc")),
                            A("⬇ Exportar Excel", href=f"/baixar-excel/{data_id}", cls="btn btn-blue", style="text-decoration:none"),
                            cls="sec-header"
                        ),
                        Div(Table(Thead(Tr(Th("Responsável"), Th("Empresa"), Th("E-mail"), Th("CPF/CNPJ"), Th("Vencimento"), Th("Status"))), Tbody(*linhas), cls="tbl"), cls="tbl-scroll"),
                        cls="sec"
                    ),
                    Div(
                        Div(
                            Div(Div("Certificados Vencidos", cls="sec-title"), Div("Envie notificações urgentes aos clientes", cls="sec-desc")),
                            Button("Enviar Todos →", cls="btn btn-red", hx_post=f"/envia-emails/{data_id}/vencido", hx_target="#res_envio_vencido", hx_indicator="#loading_vencido") if vencidos > 0 else "",
                            cls="sec-header"
                        ),
                        Div(*cards_vencidos if cards_vencidos else [Div("Nenhum certificado vencido. Tudo em dia!", cls="empty-state empty-ok")], cls="co-grid") if cards_vencidos else Div(Div("Nenhum certificado vencido. Tudo em dia!", cls="empty-state empty-ok"), style="padding:20px 24px"),
                        Div(id="res_envio_vencido", cls="mt-2", style="padding:0 24px 16px"),
                        cls="sec"
                    ),
                    Div(
                        Div(
                            Div(Div("Certificados a Vencer (30 dias)", cls="sec-title"), Div("Avisos preventivos de renovação", cls="sec-desc")),
                            Button("Enviar Todos →", cls="btn btn-amber", hx_post=f"/envia-emails/{data_id}/avencer", hx_target="#res_envio_avencer", hx_indicator="#loading_avencer") if avencer > 0 else "",
                            cls="sec-header"
                        ),
                        Div(*cards_avencer if cards_avencer else [Div("Nenhum certificado próximo de vencer", cls="empty-state empty-neutral")], cls="co-grid") if cards_avencer else Div(Div("Nenhum certificado próximo de vencer", cls="empty-state empty-neutral"), style="padding:20px 24px"),
                        Div(id="res_envio_avencer", cls="mt-2", style="padding:0 24px 16px"),
                        cls="sec"
                    ),
                    cls="dash-shell"
                ),
                cls="container"
            )
        )

    except Exception as e:
        import traceback
        print("Erro detalhado:\n", traceback.format_exc())
        return Div(
            global_css(), topbar(),
            Div(Div(Div(Div(Div("Erro ao processar", cls="sec-title"), cls="sec-header"),
                Div(P(str(e), style="color:#ef4444;font-weight:600;font-size:.9rem;margin-bottom:8px"),
                    P("Verifique se as colunas obrigatórias (CPF/CNPJ, Vencimento, Email, Empresa) estão presentes e formatadas corretamente.", style="color:#64748b;font-size:.85rem"),
                    cls="sec-body-padded"), cls="sec"), cls="dash-shell"), cls="container")
        )


@app.get("/baixar-excel/{data_id}")
def baixar_excel(request, data_id: str):
    if not is_authenticated(request):
        return RedirectResponse("/login", status_code=303)
    if f"{data_id}_excel" not in current_data:
        return Titled("Erro", Main(Section(P("Dados não encontrados"), cls="container")))
    xbytes = current_data[f"{data_id}_excel"]
    fname = f"Relatorio_Certificados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{fname}"', "X-Content-Type-Options": "nosniff"}
    return StreamingResponse(io.BytesIO(xbytes), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@app.post("/envia-emails/{data_id}/{tipo}")
async def envia_emails_lote(request, data_id: str, tipo: str):
    if not is_authenticated(request):
        return RedirectResponse("/login", status_code=303)
    df = current_data.get(data_id)
    if df is None:
        return Div(P("Dados não encontrados"), cls="sec-body-padded")

    if tipo == "vencido":
        filtro = df['Status'] == 'Vencido'
        assunto_base = "AVISO: Certificado Digital Vencido - "
        funcao_corpo = corpo_email_vencido
    elif tipo == "avencer":
        filtro = df['Status'] == 'A vencer'
        assunto_base = "ALERTA: Certificado Digital Vence em Breve - "
        funcao_corpo = corpo_email_a_vencer
    else:
        return Div(P("Tipo de e-mail inválido"), cls="sec-body-padded")

    certificados = df[filtro]
    if certificados.empty:
        return Div(P("Nenhum certificado encontrado.", style="font-weight:600;color:#2563eb"), cls="sec-body-padded")

    enviados, erros = 0, 0
    linhas = []
    for idx, cert in certificados.iterrows():
        email = (cert.get('Email') or '').strip()
        if not email or '@' not in email:
            continue
        empresa = cert.get('Empresa', 'Cliente') or 'Cliente'
        vencimento = cert.get('Vencimento', '—') or '—'
        assunto = f"{assunto_base}{empresa}"
        html = funcao_corpo(empresa, vencimento)
        ok, msg = enviar_email_smtp(email, assunto, html)
        if ok:
            enviados += 1
            linhas.append(Tr(Td("✓"), Td(empresa), Td(vencimento), Td(email), Td("Enviado")))
        else:
            erros += 1
            linhas.append(Tr(Td("✗"), Td(empresa), Td(vencimento), Td(email), Td(msg)))

    return Div(
        Div(P(Strong("Resumo: "), f"{enviados} enviados · {erros} erros · {len(certificados)} certificados",
              style="font-size:.88rem;font-weight:600;color:#0f172a;margin-bottom:12px"),
            Table(Thead(Tr(Th(""), Th("Empresa"), Th("Vencimento"), Th("E-mail"), Th("Resultado"))),
                  Tbody(*linhas)) if linhas else P("Nenhum e-mail válido encontrado.", style="color:#64748b;font-size:.88rem"),
            cls="sec-body-padded"),
        cls="sec", style="margin-top:16px"
    )


@app.post("/envia-email-individual/{data_id}/{tipo}/{idx}")
def enviar_email_individual(request, data_id: str, tipo: str, idx: str):
    if not is_authenticated(request):
        return Span("Sessão expirada. Faça login novamente.", cls="inline-feedback err")
    df = current_data.get(data_id)
    if df is None:
        return Span("Dados não encontrados", cls="inline-feedback err")
    try:
        idx = int(idx)
        cert = df.iloc[idx]
    except (ValueError, IndexError):
        return Span("Certificado não encontrado", cls="inline-feedback err")

    if tipo == "vencido":
        assunto_base = "AVISO: Certificado Digital Vencido - "
        funcao_corpo = corpo_email_vencido
    elif tipo == "avencer":
        assunto_base = "ALERTA: Certificado Digital Vence em Breve - "
        funcao_corpo = corpo_email_a_vencer
    else:
        return Span("Tipo inválido", cls="inline-feedback err")

    email = (cert.get('Email') or '').strip()
    if not email or '@' not in email:
        return Span("E-mail inválido", cls="inline-feedback err")

    empresa = cert.get('Empresa', 'Cliente') or 'Cliente'
    vencimento = cert.get('Vencimento', '—') or '—'
    assunto = f"{assunto_base}{empresa}"
    html = funcao_corpo(empresa, vencimento)
    ok, msg = enviar_email_smtp(email, assunto, html)

    if ok:
        return Span("✓ E-mail enviado!", cls="inline-feedback ok")
    else:
        return Span(f"✗ {msg}", cls="inline-feedback err")


if __name__ == "__main__":
    port = int(os.getenv("PORT", "8080"))
    serve(port=port)
